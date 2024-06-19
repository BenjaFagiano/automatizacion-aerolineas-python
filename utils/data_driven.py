import openpyxl


class DatosConfig:
    def __init__(self, destino):
        self.destino = destino

    def obtener_datos_xlsx(self, nombre_hoja):
        """
        Método para obtener los datos de la hoja de excel\n
        Ejemplo de uso:\n
        datos = DatosConfig("destino")\n
        datos.obtener_datos("Nacional")\n
        este método retorna una lista con los datos de la hoja de excel
        """
        # Cargar el archivo de excel
        rutaArchivo = f"data/datos_vuelo_{self.destino}.xlsx"
        print(rutaArchivo)
        archivo = openpyxl.load_workbook(rutaArchivo)
        hoja = archivo[nombre_hoja]
        # Obtener los datos de la hoja
        datos = []
        for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
            fila = [celda.value for celda in fila]
            datos.append(fila)
        return datos

    def obtener_datos_origen_destino_lista(self) -> list:
        """
        Método para obtener los datos de un archivo txt\n
        Ejemplo de uso:\n
        datos = DatosConfig("qa")\n
        datos.obtener_datos_txt("datos.txt")\n
        este método retorna una lista con los datos del archivo txt
        datos = [(),(),(),...]
        """
        ruta_archivo = f"data/datos_vuelos_internacional.txt"
        datos = []
        with open(ruta_archivo, 'r', encoding='utf-8') as file:
            file.readline()
            for line in file:
                # Lee cada línea y crea una tupla con los valores
                valores = line.strip().split(";")
                tupla = tuple(valores)
                datos.append(tupla)
        print(datos)
        return datos
